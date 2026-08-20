from datetime import date
from urllib.parse import quote

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, abort, jsonify, current_app
from flask_login import login_required, current_user
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from app.extensions import db
from app.models import Student, ClassRoom, AuditLog, Payment, School, AcademicYear
from app.models.user import Role
from app.students.forms import StudentForm, StudentUploadForm
from app.utils.decorators import write_access_required, roles_required
from app.utils.helpers import scope_query_to_school, current_school_id, is_super_admin
from app.utils.db_safety import safe_commit
from app.services.export_service import export_excel
from app.services.storage_service import save_image, StorageError

students_bp = Blueprint("students", __name__, template_folder="../templates/students")


def _populate_class_choices(form, school_id):
    q = ClassRoom.query
    if school_id is not None:
        q = q.filter_by(school_id=school_id)
    form.class_id.choices = [(0, "-- No Class --")] + [(c.id, c.name) for c in q.order_by(ClassRoom.name).all()]


@students_bp.route("/")
@login_required
def list_students():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("q", "").strip()
    status = request.args.get("status", "")
    class_id = request.args.get("class_id", type=int)

    query = scope_query_to_school(Student.query, Student)
    if search:
        like = f"%{search}%"
        query = query.filter(
            (Student.first_name.ilike(like))
            | (Student.last_name.ilike(like))
            | (Student.student_id.ilike(like))
            | (Student.guardian_name.ilike(like))
            | (Student.guardian_contact.ilike(like))
        )
    if status:
        query = query.filter(Student.status == status)
    if class_id:
        query = query.filter(Student.class_id == class_id)

    pagination = query.order_by(Student.created_at.desc()).paginate(page=page, per_page=20, error_out=False)

    classes_q = ClassRoom.query
    if not is_super_admin():
        classes_q = classes_q.filter_by(school_id=current_school_id())
    classes = classes_q.order_by(ClassRoom.name).all()

    return render_template(
        "students/list.html", students=pagination.items, pagination=pagination,
        search=search, status=status, class_id=class_id, classes=classes,
    )


@students_bp.route("/<int:student_id>")
@login_required
def view_student(student_id):
    student = Student.query.get_or_404(student_id)
    if not is_super_admin() and student.school_id != current_school_id():
        abort(403)
    payments = student.payments.order_by(Payment.payment_date.desc()).limit(50).all()
    return render_template("students/view.html", student=student, payments=payments)


@students_bp.route("/create", methods=["GET", "POST"])
@login_required
@roles_required(Role.SUPER_ADMIN, Role.SCHOOL_ADMIN)
def create_student():
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school context before adding a student.", "warning")
        return redirect(url_for("dashboard.index"))

    form = StudentForm()
    _populate_class_choices(form, school_id)

    if form.validate_on_submit():
        try:
            photo_path = save_image(form.photo.data, subfolder="students")
        except StorageError as exc:
            # Previously this had no try/except at all, so any storage
            # failure (bad path, disk permissions, corrupt/oversized image)
            # crashed straight to a raw 500. Now the user gets the exact
            # reason and can fix it (or retry) without losing their other
            # form input.
            form.photo.errors.append(exc.user_message)
            return render_template("students/form.html", form=form, title="Add Student")

        student = Student(
            school_id=school_id,
            student_id=Student.generate_student_id(school_id),
            first_name=form.first_name.data.strip(),
            last_name=form.last_name.data.strip(),
            gender=form.gender.data,
            date_of_birth=form.date_of_birth.data,
            guardian_name=form.guardian_name.data.strip(),
            guardian_contact=form.guardian_contact.data.strip(),
            class_id=form.class_id.data or None,
            admission_date=form.admission_date.data or date.today(),
            status=form.status.data,
            photo=photo_path,
        )
        db.session.add(student)
        if safe_commit(log_context=f"create_student school={school_id}"):
            AuditLog.log(
                "student_created", description=f"Student {student.student_id} created", entity_type="student",
                entity_id=student.id, user=current_user, school_id=school_id,
            )
            flash(f"Student {student.full_name} ({student.student_id}) created successfully.", "success")
            return redirect(url_for("students.list_students"))

    return render_template("students/form.html", form=form, title="Add Student")


@students_bp.route("/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required(Role.SUPER_ADMIN, Role.SCHOOL_ADMIN)
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    if not is_super_admin() and student.school_id != current_school_id():
        abort(403)

    form = StudentForm(obj=student)
    _populate_class_choices(form, student.school_id)
    if request.method == "GET":
        form.class_id.data = student.class_id or 0

    if form.validate_on_submit():
        student.first_name = form.first_name.data.strip()
        student.last_name = form.last_name.data.strip()
        student.gender = form.gender.data
        student.date_of_birth = form.date_of_birth.data
        student.guardian_name = form.guardian_name.data.strip()
        student.guardian_contact = form.guardian_contact.data.strip()
        student.class_id = form.class_id.data or None
        student.admission_date = form.admission_date.data
        student.status = form.status.data
        if isinstance(form.photo.data, FileStorage) and form.photo.data.filename:
            try:
                student.photo = save_image(form.photo.data, subfolder="students", old_reference=student.photo)
            except StorageError as exc:
                form.photo.errors.append(exc.user_message)
                return render_template("students/form.html", form=form, title="Edit Student", student=student)
        if safe_commit(log_context=f"edit_student {student.id}"):
            AuditLog.log(
                "student_updated", description=f"Student {student.student_id} updated", entity_type="student",
                entity_id=student.id, user=current_user,
            )
            flash("Student updated successfully.", "success")
            return redirect(url_for("students.view_student", student_id=student.id))

    return render_template("students/form.html", form=form, title="Edit Student", student=student)


@students_bp.route("/<int:student_id>/delete", methods=["POST"])
@login_required
@write_access_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    if not is_super_admin() and student.school_id != current_school_id():
        abort(403)
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN):
        abort(403)

    student_ref = f"{student.student_id} - {student.full_name}"
    old_photo = student.photo
    db.session.delete(student)
    if safe_commit(
        friendly_message="This student can't be deleted because they have payment records. "
                          "Deactivate the student instead to preserve financial history.",
        log_context=f"delete_student {student_id}",
    ):
        if old_photo:
            from app.services.storage_service import delete_file
            delete_file(old_photo)
        AuditLog.log("student_deleted", description=f"Student {student_ref} deleted", user=current_user)
        flash("Student deleted.", "info")
    return redirect(url_for("students.list_students"))


@students_bp.route("/<int:student_id>/deactivate", methods=["POST"])
@login_required
@roles_required(Role.SUPER_ADMIN, Role.SCHOOL_ADMIN)
def deactivate_student(student_id):
    student = Student.query.get_or_404(student_id)
    if not is_super_admin() and student.school_id != current_school_id():
        abort(403)
    student.status = "deactivated"
    if safe_commit(log_context=f"deactivate_student {student_id}"):
        AuditLog.log(
            "student_deactivated", description=f"Student {student.student_id} deactivated",
            entity_type="student", entity_id=student.id, user=current_user,
        )
        flash("Student deactivated.", "info")
    return redirect(url_for("students.view_student", student_id=student.id))


@students_bp.route("/upload", methods=["GET", "POST"])
@login_required
@roles_required(Role.SUPER_ADMIN, Role.SCHOOL_ADMIN)
def upload_students():
    school_id = current_school_id()
    if school_id is None and not is_super_admin():
        flash("No school context found.", "warning")
        return redirect(url_for("students.list_students"))

    form = StudentUploadForm()
    if form.validate_on_submit():
        wb = load_workbook(form.file.data, data_only=True)
        ws = wb.active
        created = 0
        errors = []
        header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        required_cols = {"first_name", "last_name"}
        if not required_cols.issubset(set(header)):
            flash("Excel file must contain at least 'first_name' and 'last_name' columns.", "danger")
            return redirect(url_for("students.upload_students"))

        col_index = {name: idx for idx, name in enumerate(header)}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[col_index.get("first_name", 0)]:
                continue
            # Each row gets its own savepoint so one bad row can't poison the
            # whole transaction and silently drop every subsequent row - a
            # real risk on PostgreSQL, where a failed statement aborts the
            # entire transaction until it's rolled back.
            try:
                with db.session.begin_nested():
                    first_name = str(row[col_index["first_name"]]).strip()
                    last_name = str(row[col_index.get("last_name", 1)]).strip()
                    gender = str(row[col_index["gender"]]).strip().lower() if "gender" in col_index and row[col_index["gender"]] else "male"
                    guardian_name = str(row[col_index["guardian_name"]]).strip() if "guardian_name" in col_index and row[col_index["guardian_name"]] else ""
                    guardian_contact = str(row[col_index["guardian_contact"]]).strip() if "guardian_contact" in col_index and row[col_index["guardian_contact"]] else ""

                    student = Student(
                        school_id=school_id,
                        student_id=Student.generate_student_id(school_id),
                        first_name=first_name,
                        last_name=last_name,
                        gender=gender,
                        guardian_name=guardian_name,
                        guardian_contact=guardian_contact,
                        admission_date=date.today(),
                        status="active",
                    )
                    db.session.add(student)
                    db.session.flush()
                created += 1
            except Exception as exc:  # noqa: BLE001
                current_app.logger.warning("Row %s skipped during student import: %s", row_num, exc)
                errors.append(f"Row {row_num}: {exc}")

        if safe_commit(log_context=f"upload_students school={school_id}"):
            AuditLog.log(
                "students_bulk_uploaded", description=f"{created} students uploaded via Excel",
                user=current_user, school_id=school_id,
            )
            flash(f"{created} students imported successfully." + (f" {len(errors)} rows skipped." if errors else ""), "success")
            return redirect(url_for("students.list_students"))

    return render_template("students/upload.html", form=form)


@students_bp.route("/download")
@login_required
def download_students():
    query = scope_query_to_school(Student.query, Student)
    students = query.order_by(Student.student_id).all()
    headers = ["Student ID", "First Name", "Last Name", "Gender", "Guardian", "Guardian Contact", "Class", "Status", "Admission Date"]
    rows = [
        [
            s.student_id, s.first_name, s.last_name, s.gender, s.guardian_name, s.guardian_contact,
            s.classroom.name if s.classroom else "", s.status, s.admission_date.isoformat() if s.admission_date else "",
        ]
        for s in students
    ]
    mem = export_excel(headers, rows, sheet_title="Students")
    return send_file(mem, as_attachment=True, download_name="students.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------- Live AJAX search (used by the payment collection screen) ----------

@students_bp.route("/api/search")
@login_required
def api_search_students():
    """JSON endpoint for the live student search used on the "Record Payment"
    screen. Always scoped to the current user's school (or, for a Super
    Admin, to the ?school_id they've selected) - never returns cross-tenant
    results."""
    q = request.args.get("q", "").strip()
    school_id = current_school_id()

    if not q or len(q) < 2:
        return jsonify({"results": []})

    query = Student.query.filter(Student.status == "active")
    if school_id is not None:
        query = query.filter(Student.school_id == school_id)
    elif not is_super_admin():
        return jsonify({"results": []})

    like = f"%{q}%"

    # also allow matching by class name
    query = query.outerjoin(ClassRoom, Student.class_id == ClassRoom.id).filter(
        (Student.student_id.ilike(like))
        | (Student.first_name.ilike(like))
        | (Student.last_name.ilike(like))
        | (Student.guardian_name.ilike(like))
        | (Student.guardian_contact.ilike(like))
        | (ClassRoom.name.ilike(like))
    )

    students = query.order_by(Student.first_name).limit(15).all()
    return jsonify({
        "results": [
            {
                "id": s.id,
                "student_id": s.student_id,
                "name": s.full_name,
                "guardian_name": s.guardian_name,
                "guardian_contact": s.guardian_contact,
                "class_name": s.classroom.name if s.classroom else "",
                "balance": s.total_paid,
            }
            for s in students
        ]
    })


# ---------- Students Owing / Reminders ----------

@students_bp.route("/owing")
@login_required
def students_owing():
    """Searchable, filterable dashboard of active students and their
    contribution status, with one-click WhatsApp reminders and a direct
    "Record Payment" action - no need to navigate through multiple pages.
    Available to admins, accountants, and collectors (the roles responsible
    for chasing payments); never exposes school-wide balance/financial
    totals to Collectors (see list_payments/list for that restriction)."""
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN, Role.ACCOUNTANT, Role.COLLECTOR):
        abort(403)

    school_id = current_school_id()
    expected_amount = request.args.get("expected", type=float) or 0.0
    search = request.args.get("q", "").strip()
    class_id = request.args.get("class_id", type=int)
    status_filter = request.args.get("status", "")  # '', 'owing', 'paid'
    min_owed = request.args.get("min_owed", type=float)

    query = Student.query.filter(Student.status == "active")
    if school_id is not None:
        query = query.filter(Student.school_id == school_id)
    if class_id:
        query = query.filter(Student.class_id == class_id)
    if search:
        like = f"%{search}%"
        query = query.outerjoin(ClassRoom, Student.class_id == ClassRoom.id).filter(
            (Student.first_name.ilike(like)) | (Student.last_name.ilike(like))
            | (Student.student_id.ilike(like)) | (Student.guardian_name.ilike(like))
            | (Student.guardian_contact.ilike(like)) | (ClassRoom.name.ilike(like))
        )
    students = query.order_by(Student.first_name).all()

    school = School.query.get(school_id) if school_id else None
    rows = []
    for s in students:
        paid = s.total_paid
        balance = max(expected_amount - paid, 0) if expected_amount else 0
        is_owing = bool(expected_amount) and balance > 0

        if status_filter == "owing" and not is_owing:
            continue
        if status_filter == "paid" and (is_owing or not expected_amount):
            continue
        if min_owed is not None and balance < min_owed:
            continue

        last_payment = (
            s.payments.filter(Payment.is_void.is_(False)).order_by(Payment.payment_date.desc()).first()
        )

        message = (
            f"Assalamu Alaikum. This is a reminder from {school.name if school else 'the Madrassa'}. "
            f"Your child {s.first_name} has an outstanding balance"
            + (f" of GH₵{balance:.2f}." if expected_amount else ".")
            + " Kindly make payment at your earliest convenience. JazakAllahu Khairan."
        )
        whatsapp_number = "".join(ch for ch in (s.guardian_contact or "") if ch.isdigit())
        whatsapp_url = f"https://wa.me/{whatsapp_number}?text={quote(message)}" if whatsapp_number else None

        rows.append({
            "student": s,
            "paid": paid,
            "balance": balance,
            "is_owing": is_owing,
            "last_payment": last_payment,
            "whatsapp_url": whatsapp_url,
            "has_contact": bool(whatsapp_number),
        })

    classes_q = ClassRoom.query
    if school_id is not None:
        classes_q = classes_q.filter_by(school_id=school_id)
    classes = classes_q.order_by(ClassRoom.name).all()

    can_record_payment = current_user.role != Role.TEACHER

    return render_template(
        "students/owing.html", rows=rows, expected_amount=expected_amount,
        schools=School.query.order_by(School.name).all() if is_super_admin() else [],
        selected_school_id=school_id, search=search, classes=classes, class_id=class_id,
        status_filter=status_filter, min_owed=min_owed, can_record_payment=can_record_payment,
    )


# ---------------------------------------------------------------------------
# Student Promotion
# ---------------------------------------------------------------------------

@students_bp.route('/promote', methods=['GET', 'POST'])
@login_required
def promote_students():
    if current_user.role not in [Role.SUPER_ADMIN, Role.SCHOOL_ADMIN]:
        flash("Permission denied.", "danger")
        return redirect(url_for("dashboard.index"))
    school_id = current_school_id()
    classes = ClassRoom.query.filter_by(school_id=school_id, is_active=True).order_by(ClassRoom.name).all()
    years = AcademicYear.query.filter_by(school_id=school_id, is_active=True).order_by(AcademicYear.start_date.desc()).all()

    if request.method == "POST":
        from_class_id = request.form.get("from_class_id", type=int)
        to_class_id = request.form.get("to_class_id", type=int)
        year_id = request.form.get("academic_year_id", type=int)
        action = request.form.get("action")

        if not from_class_id or not to_class_id or not year_id:
            flash("Please select all required fields.", "danger")
            return redirect(request.url)

        student_ids = request.form.getlist("student_ids")
        if not student_ids:
            flash("No students selected.", "warning")
            return redirect(request.url)

        count = 0
        for sid in student_ids:
            student = Student.query.filter_by(id=sid, school_id=school_id, class_id=from_class_id).first()
            if not student:
                continue
            if action == "promote":
                student.class_id = to_class_id
                student.academic_year_id = year_id
                student.promotion_status = "promoted"
            elif action == "repeat":
                student.promotion_status = "repeated"
            elif action == "graduate":
                student.status = "graduated"
                student.promotion_status = "graduated"
            count += 1

        if safe_commit():
            flash(f"{count} students {action}d successfully.", "success")
        return redirect(url_for("students.list_students"))

    return render_template("students/promote.html", classes=classes, years=years)


@students_bp.route('/promote/load-students', methods=['POST'])
@login_required
def load_students_for_promotion():
    if current_user.role not in [Role.SUPER_ADMIN, Role.SCHOOL_ADMIN]:
        return ""
    school_id = current_school_id()
    class_id = request.form.get("class_id", type=int)
    students = Student.query.filter_by(school_id=school_id, class_id=class_id, status="active").order_by(Student.first_name).all()
    return render_template("students/_promotion_students.html", students=students)


# ---------------------------------------------------------------------------
# Student ID Card
# ---------------------------------------------------------------------------

@students_bp.route('/<int:id>/id-card')
@login_required
def student_id_card(id):
    school_id = current_school_id()
    student = Student.query.filter_by(id=id, school_id=school_id).first_or_404()
    school = School.query.get(school_id)
    from app.services.export_service import generate_student_id_card_pdf
    pdf = generate_student_id_card_pdf(student, school)
    return send_file(pdf, download_name=f"id_card_{student.student_id}.pdf", as_attachment=True)