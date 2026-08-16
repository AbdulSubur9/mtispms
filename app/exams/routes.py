from datetime import date
import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import ClassRoom, Subject, Exam, ExamSubject, Result, GradingScaleBand, AuditLog, Student
from app.models.user import Role
from app.exams.forms import SubjectForm, ExamForm, GradingBandForm
from app.utils.helpers import current_school_id, is_super_admin
from app.utils.db_safety import safe_commit
from app.services.results_service import compute_exam_summary, ordinal
from app.services.export_service import generate_student_report_pdf, generate_class_result_sheet_pdf
from openpyxl import load_workbook, Workbook

exams_bp = Blueprint("exams", __name__, template_folder="../templates/exams")

ACADEMIC_STAFF = (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN, Role.TEACHER)


def _assert_can_manage_class(classroom):
    """Same tenant + teacher-ownership rule used for attendance: a Teacher
    may only work with classes assigned to them; School Admin/Super Admin
    may work with any class in their own school."""
    if not is_super_admin() and classroom.school_id != current_school_id():
        abort(403)
    if current_user.role == Role.TEACHER and classroom.teacher_id != current_user.id:
        abort(403)


def _teacher_classes():
    query = ClassRoom.query
    if current_user.role == Role.TEACHER:
        query = query.filter(ClassRoom.teacher_id == current_user.id)
    elif not is_super_admin():
        query = query.filter(ClassRoom.school_id == current_school_id())
    return query.order_by(ClassRoom.name).all()


def _ensure_default_grading_scale(school_id):
    """A school with no grading scale configured yet gets the sensible
    A-F default from the spec, seeded once, and fully editable afterwards -
    grading is never hard-coded into the app itself."""
    if GradingScaleBand.query.filter_by(school_id=school_id).count() == 0:
        for band in GradingScaleBand.default_scale_for(school_id):
            db.session.add(band)
        db.session.commit()


# ---------------------------------------------------------------------------
# Subjects
# ---------------------------------------------------------------------------

@exams_bp.route("/subjects")
@login_required
def list_subjects():
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    school_id = current_school_id()
    query = Subject.query
    if school_id is not None:
        query = query.filter_by(school_id=school_id)
    subjects = query.order_by(Subject.name).all()
    return render_template("exams/subjects.html", subjects=subjects)


@exams_bp.route("/subjects/create", methods=["GET", "POST"])
@login_required
def create_subject():
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN, Role.TEACHER):
        abort(403)
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("exams.list_subjects"))

    form = SubjectForm()
    if form.validate_on_submit():
        subject = Subject(school_id=school_id, name=form.name.data.strip(), is_active=form.is_active.data)
        db.session.add(subject)
        if safe_commit(
            friendly_message="A subject with this name already exists for this school.",
            log_context=f"create_subject school={school_id}",
        ):
            flash("Subject added.", "success")
            return redirect(url_for("exams.list_subjects"))
    return render_template("exams/subject_form.html", form=form, title="Add Subject")


@exams_bp.route("/subjects/<int:subject_id>/toggle", methods=["POST"])
@login_required
def toggle_subject(subject_id):
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN):
        abort(403)
    subject = Subject.query.get_or_404(subject_id)
    if not is_super_admin() and subject.school_id != current_school_id():
        abort(403)
    subject.is_active = not subject.is_active
    if safe_commit(log_context=f"toggle_subject {subject_id}"):
        flash(f"Subject {'activated' if subject.is_active else 'deactivated'}.", "info")
    return redirect(url_for("exams.list_subjects"))


# ---------------------------------------------------------------------------
# Grading scale
# ---------------------------------------------------------------------------

@exams_bp.route("/grading")
@login_required
def grading_scale():
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN):
        abort(403)
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("dashboard.index"))
    _ensure_default_grading_scale(school_id)
    bands = GradingScaleBand.query.filter_by(school_id=school_id).order_by(GradingScaleBand.min_percentage.desc()).all()
    return render_template("exams/grading.html", bands=bands)


@exams_bp.route("/grading/create", methods=["GET", "POST"])
@login_required
def create_grading_band():
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN):
        abort(403)
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("exams.grading_scale"))

    form = GradingBandForm()
    if form.validate_on_submit():
        band = GradingScaleBand(
            school_id=school_id, grade=form.grade.data.strip().upper(),
            min_percentage=form.min_percentage.data, max_percentage=form.max_percentage.data,
            remark=form.remark.data,
        )
        db.session.add(band)
        if safe_commit(log_context=f"create_grading_band school={school_id}"):
            AuditLog.log(
                "settings_changed", description=f"Grading band {band.grade} added", user=current_user,
                school_id=school_id,
            )
            flash("Grading band added.", "success")
            return redirect(url_for("exams.grading_scale"))
    return render_template("exams/grading_form.html", form=form, title="Add Grading Band")


@exams_bp.route("/grading/<int:band_id>/delete", methods=["POST"])
@login_required
def delete_grading_band(band_id):
    if current_user.role not in (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN):
        abort(403)
    band = GradingScaleBand.query.get_or_404(band_id)
    if not is_super_admin() and band.school_id != current_school_id():
        abort(403)
    db.session.delete(band)
    if safe_commit(log_context=f"delete_grading_band {band_id}"):
        flash("Grading band removed.", "info")
    return redirect(url_for("exams.grading_scale"))


# ---------------------------------------------------------------------------
# Exams
# ---------------------------------------------------------------------------

@exams_bp.route("/")
@login_required
def list_exams():
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    classes = _teacher_classes()
    class_ids = [c.id for c in classes]
    if not class_ids:
        exams = []
    else:
        exams = Exam.query.filter(Exam.class_id.in_(class_ids)).order_by(Exam.created_at.desc()).all()
    return render_template("exams/list.html", exams=exams)


@exams_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_exam():
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("exams.list_exams"))

    form = ExamForm()
    classes = _teacher_classes()
    form.class_id.choices = [(c.id, c.name) for c in classes]

    subjects = Subject.query.filter_by(school_id=school_id, is_active=True).order_by(Subject.name).all()
    if not subjects:
        flash("Create at least one subject before creating an examination.", "warning")
        return redirect(url_for("exams.create_subject"))

    if form.validate_on_submit():
        classroom = ClassRoom.query.get_or_404(form.class_id.data)
        _assert_can_manage_class(classroom)

        selected_subject_ids = request.form.getlist("subject_ids")
        if not selected_subject_ids:
            flash("Select at least one subject for this examination.", "danger")
            return render_template("exams/exam_form.html", form=form, subjects=subjects, title="Create Examination")

        exam = Exam(
            school_id=school_id, class_id=classroom.id, created_by_id=current_user.id,
            name=form.name.data.strip(), exam_date=form.exam_date.data,
        )
        db.session.add(exam)
        db.session.flush()

        for subject_id in selected_subject_ids:
            max_marks = request.form.get(f"max_marks_{subject_id}", type=float) or 100
            db.session.add(ExamSubject(exam_id=exam.id, subject_id=int(subject_id), max_marks=max_marks))

        if safe_commit(log_context=f"create_exam school={school_id}"):
            AuditLog.log(
                "exam_created", description=f"Exam '{exam.name}' created for {classroom.name}",
                entity_type="exam", entity_id=exam.id, user=current_user, school_id=school_id,
            )
            flash("Examination created. You can now enter marks for each subject.", "success")
            return redirect(url_for("exams.view_exam", exam_id=exam.id))

    return render_template("exams/exam_form.html", form=form, subjects=subjects, title="Create Examination")


@exams_bp.route("/<int:exam_id>")
@login_required
def view_exam(exam_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    exam = Exam.query.get_or_404(exam_id)
    _assert_can_manage_class(exam.classroom)

    _ensure_default_grading_scale(exam.school_id)
    summary = compute_exam_summary(exam)
    return render_template("exams/view.html", exam=exam, summary=summary, ordinal=ordinal)


@exams_bp.route("/<int:exam_id>/enter/<int:exam_subject_id>", methods=["GET", "POST"])
@login_required
def enter_marks(exam_id, exam_subject_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    exam = Exam.query.get_or_404(exam_id)
    _assert_can_manage_class(exam.classroom)
    exam_subject = ExamSubject.query.get_or_404(exam_subject_id)
    if exam_subject.exam_id != exam.id:
        abort(404)

    students = exam.classroom.students.filter_by(status="active").order_by(Student.first_name).all()
    existing = {r.student_id: r for r in exam_subject.results.all()}

    if request.method == "POST":
        for student in students:
            raw = request.form.get(f"marks_{student.id}", "").strip()
            comment = request.form.get(f"comment_{student.id}", "").strip()
            if raw == "":
                continue
            try:
                marks = float(raw)
            except ValueError:
                continue
            marks = max(0, min(marks, float(exam_subject.max_marks)))  # clamp to valid range

            record = existing.get(student.id)
            if record:
                record.marks_obtained = marks
                record.teacher_comment = comment or None
            else:
                db.session.add(Result(
                    school_id=exam.school_id, exam_subject_id=exam_subject.id, student_id=student.id,
                    recorded_by_id=current_user.id, marks_obtained=marks, teacher_comment=comment or None,
                ))

        if safe_commit(log_context=f"enter_marks exam_subject={exam_subject_id}"):
            AuditLog.log(
                "result_uploaded", description=f"Marks entered for {exam_subject.subject.name} in {exam.name}",
                entity_type="exam", entity_id=exam.id, user=current_user, school_id=exam.school_id,
            )
            flash("Marks saved.", "success")
            return redirect(url_for("exams.enter_marks", exam_id=exam.id, exam_subject_id=exam_subject.id))

    return render_template(
        "exams/enter_marks.html", exam=exam, exam_subject=exam_subject, students=students, existing=existing
    )


@exams_bp.route("/<int:exam_id>/publish", methods=["POST"])
@login_required
def publish_exam(exam_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    exam = Exam.query.get_or_404(exam_id)
    _assert_can_manage_class(exam.classroom)
    exam.is_published = not exam.is_published
    if safe_commit(log_context=f"publish_exam {exam_id}"):
        AuditLog.log(
            "settings_changed", description=f"Exam '{exam.name}' {'published' if exam.is_published else 'unpublished'}",
            entity_type="exam", entity_id=exam.id, user=current_user,
        )
        flash(f"Exam {'published' if exam.is_published else 'unpublished'}.", "info")
    return redirect(url_for("exams.view_exam", exam_id=exam.id))


@exams_bp.route("/<int:exam_id>/student/<int:student_id>/report")
@login_required
def student_report(exam_id, student_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    exam = Exam.query.get_or_404(exam_id)
    _assert_can_manage_class(exam.classroom)
    student = Student.query.get_or_404(student_id)
    if student.school_id != exam.school_id:
        abort(403)

    summary = compute_exam_summary(exam)
    row = next((r for r in summary if r["student"].id == student.id), None)
    if row is None:
        abort(404)

    mem = generate_student_report_pdf(exam.classroom.school, exam, row, ordinal)
    return send_file(mem, as_attachment=True, download_name=f"{student.student_id}_{exam.name}.pdf".replace(" ", "_"),
                      mimetype="application/pdf")


@exams_bp.route("/<int:exam_id>/result-sheet")
@login_required
def class_result_sheet(exam_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    exam = Exam.query.get_or_404(exam_id)
    _assert_can_manage_class(exam.classroom)

    summary = compute_exam_summary(exam)
    mem = generate_class_result_sheet_pdf(exam.classroom.school, exam, summary, ordinal)
    return send_file(mem, as_attachment=True, download_name=f"{exam.name}_result_sheet.pdf".replace(" ", "_"),
                      mimetype="application/pdf")


# ---------------------------------------------------------------------------
# Bulk Marks Import
# ---------------------------------------------------------------------------

@exams_bp.route("/<int:exam_id>/import-marks", methods=["GET", "POST"])
@login_required
def import_marks(exam_id):
    if current_user.role not in ACADEMIC_STAFF:
        abort(403)
    school_id = current_school_id()
    exam = Exam.query.filter_by(id=exam_id, school_id=school_id).first_or_404()
    if exam.status == "locked":
        flash("Cannot import marks for a locked exam.", "danger")
        return redirect(url_for("exams.view_exam", exam_id=exam.id))

    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith((".xlsx", ".xls")):
            flash("Please upload a valid Excel file (.xlsx)", "danger")
            return redirect(request.url)

        try:
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            required = ["Student ID", "Subject", "Marks"]
            if not all(h in headers for h in required):
                flash(f"Excel must have columns: {', '.join(required)}", "danger")
                return redirect(request.url)

            sid_idx = headers.index("Student ID")
            subj_idx = headers.index("Subject")
            marks_idx = headers.index("Marks")

            imported = 0
            errors = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                student_id_str = str(row[sid_idx]) if row[sid_idx] else None
                subject_name = str(row[subj_idx]) if row[subj_idx] else None
                marks = row[marks_idx]

                if not student_id_str or not subject_name or marks is None:
                    continue

                student = Student.query.filter_by(school_id=school_id, student_id=student_id_str).first()
                if not student:
                    errors.append(f"Student '{student_id_str}' not found")
                    continue

                subject = Subject.query.filter_by(school_id=school_id, name=subject_name).first()
                if not subject:
                    errors.append(f"Subject '{subject_name}' not found")
                    continue

                exam_subj = ExamSubject.query.filter_by(exam_id=exam.id, subject_id=subject.id).first()
                if not exam_subj:
                    errors.append(f"Subject '{subject_name}' not in this exam")
                    continue

                try:
                    marks_val = float(marks)
                    if marks_val < 0 or marks_val > float(exam_subj.max_marks):
                        errors.append(f"Invalid marks for {student_id_str} in {subject_name}")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Non-numeric marks for {student_id_str}")
                    continue

                result = Result.query.filter_by(exam_subject_id=exam_subj.id, student_id=student.id).first()
                if result:
                    result.marks_obtained = marks_val
                    result.recorded_by_id = current_user.id
                else:
                    result = Result(
                        school_id=school_id,
                        exam_subject_id=exam_subj.id,
                        student_id=student.id,
                        marks_obtained=marks_val,
                        recorded_by_id=current_user.id,
                    )
                    db.session.add(result)
                imported += 1

            if safe_commit():
                flash(f"Imported {imported} marks. {len(errors)} errors.", "success" if not errors else "warning")
                for err in errors[:10]:
                    flash(err, "warning")
        except Exception as e:
            flash(f"Import failed: {str(e)}", "danger")
        return redirect(url_for("exams.view_exam", exam_id=exam.id))

    return render_template("exams/import_marks.html", exam=exam)


@exams_bp.route("/<int:exam_id>/import-template")
@login_required
def download_import_template(exam_id):
    school_id = current_school_id()
    exam = Exam.query.filter_by(id=exam_id, school_id=school_id).first_or_404()
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Import"
    ws.append(["Student ID", "Student Name", "Subject", "Marks", "Max Marks"])

    for es in exam.exam_subjects:
        for student in exam.classroom.students.filter_by(status="active").order_by(Student.first_name):
            ws.append([student.student_id, student.full_name, es.subject.name, "", str(es.max_marks)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"marks_template_{exam.id}.xlsx", as_attachment=True)