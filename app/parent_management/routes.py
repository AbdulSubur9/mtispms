from datetime import datetime, date
from urllib.parse import quote

from flask import render_template, redirect, url_for, flash, request, send_file, abort
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.parent_management import parent_management_bp
from app.parent_management.forms import ParentCreateForm, ParentImportForm
from app.parent_management.credentials import generate_temp_password, generate_unique_username
from app.models import Parent, ParentStudent, Student, User, ClassRoom, School, AuditLog
from app.models.user import Role
from app.utils.decorators import roles_required
from app.utils.helpers import current_school_id, is_super_admin
from app.utils.db_safety import safe_commit

ADMIN_ROLES = (Role.SUPER_ADMIN, Role.SCHOOL_ADMIN)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@parent_management_bp.route("/")
@login_required
@roles_required(*ADMIN_ROLES)
def dashboard():
    school_id = current_school_id()

    parent_q = Parent.query.join(User, Parent.user_id == User.id)
    if school_id is not None:
        parent_q = parent_q.filter(Parent.school_id == school_id)

    total_parents = parent_q.count()
    active_parents = parent_q.filter(User.is_active_user.is_(True)).count()

    all_parents = parent_q.all()
    linked_parent_ids = {link.parent_id for link in ParentStudent.query.filter(
        ParentStudent.parent_id.in_([p.id for p in all_parents])
    ).all()} if all_parents else set()
    parents_without_students = sum(1 for p in all_parents if p.id not in linked_parent_ids)

    student_q = Student.query.filter(Student.status == "active")
    if school_id is not None:
        student_q = student_q.filter(Student.school_id == school_id)
    all_students = student_q.all()
    linked_student_ids = {link.student_id for link in ParentStudent.query.filter(
        ParentStudent.student_id.in_([s.id for s in all_students])
    ).all()} if all_students else set()
    students_without_parent = sum(1 for s in all_students if s.id not in linked_student_ids)

    recent_parents = parent_q.order_by(Parent.created_at.desc()).limit(8).all()

    return render_template(
        "parent_management/dashboard.html",
        total_parents=total_parents, active_parents=active_parents,
        parents_without_students=parents_without_students, students_without_parent=students_without_parent,
        recent_parents=recent_parents,
    )


# ---------------------------------------------------------------------------
# List / search
# ---------------------------------------------------------------------------

@parent_management_bp.route("/list")
@login_required
@roles_required(*ADMIN_ROLES)
def list_parents():
    school_id = current_school_id()
    page = request.args.get("page", 1, type=int)
    search = request.args.get("q", "").strip()

    query = Parent.query.join(User, Parent.user_id == User.id)
    if school_id is not None:
        query = query.filter(Parent.school_id == school_id)

    if search:
        like = f"%{search}%"
        matching_students = Student.query.filter(
            (Student.first_name.ilike(like)) | (Student.last_name.ilike(like)) | (Student.student_id.ilike(like))
        ).all()
        matching_student_ids = [s.id for s in matching_students]
        parent_ids_via_student = [
            link.parent_id for link in ParentStudent.query.filter(
                ParentStudent.student_id.in_(matching_student_ids)
            ).all()
        ] if matching_student_ids else []

        conditions = [
            User.first_name.ilike(like), User.last_name.ilike(like),
            User.email.ilike(like), User.phone.ilike(like),
        ]
        if parent_ids_via_student:
            conditions.append(Parent.id.in_(parent_ids_via_student))

        from sqlalchemy import or_
        query = query.filter(or_(*conditions))

    pagination = query.order_by(User.first_name).paginate(page=page, per_page=20, error_out=False)
    return render_template("parent_management/list.html", parents=pagination.items, pagination=pagination, search=search)


# ---------------------------------------------------------------------------
# Create parent (+ credential generation)
# ---------------------------------------------------------------------------

@parent_management_bp.route("/create", methods=["GET", "POST"])
@login_required
@roles_required(*ADMIN_ROLES)
def create_parent():
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("parent_management.dashboard"))

    form = ParentCreateForm()
    if form.validate_on_submit():
        username = generate_unique_username("parent")
        temp_password = generate_temp_password()

        user = User(
            username=username, email=form.email.data.strip().lower(),
            first_name=form.first_name.data.strip(), last_name=form.last_name.data.strip(),
            phone=form.phone.data.strip(), role=Role.PARENT, school_id=school_id,
            is_active_user=True, must_change_password=True,
        )
        user.set_password(temp_password)
        db.session.add(user)
        db.session.flush()

        parent = Parent(
            school_id=school_id, user_id=user.id, relationship=form.relationship.data,
            occupation=form.occupation.data, address=form.address.data,
            alternative_phone=form.alternative_phone.data,
            emergency_contact=form.emergency_contact.data, emergency_contact_name=form.emergency_contact_name.data,
        )
        db.session.add(parent)

        if not safe_commit(log_context=f"create_parent school={school_id}"):
            return render_template("parent_management/create.html", form=form)

        AuditLog.log(
            "user_created", description=f"Parent account created for {user.full_name}",
            entity_type="parent", entity_id=parent.id, user=current_user, school_id=school_id,
        )
        flash(f"Parent account created for {user.full_name}.", "success")
        # Never persist or log the plaintext password - hand it to the
        # admin ONCE, on this single response, via a one-time reveal page.
        return render_template(
            "parent_management/credentials_created.html",
            parent=parent, username=username, temp_password=temp_password,
        )

    return render_template("parent_management/create.html", form=form)


@parent_management_bp.route("/<int:parent_id>/whatsapp-message")
@login_required
@roles_required(*ADMIN_ROLES)
def whatsapp_credentials_message(parent_id):
    """Builds a wa.me share link with the account details, for the ONE-TIME
    reveal page only - never reconstructable afterward, since the
    plaintext password is never stored. If a real WhatsApp Business API
    integration isn't configured, this is the honest fallback: a
    copy/share link the admin sends themselves, not a fake "sent"
    confirmation."""
    parent = Parent.query.get_or_404(parent_id)
    if not is_super_admin() and parent.school_id != current_school_id():
        abort(403)
    username = request.args.get("username", "")
    temp_password = request.args.get("password", "")
    school = School.query.get(parent.school_id)

    message = (
        f"Assalamu Alaikum {parent.full_name},\n\n"
        f"An account has been created for you on {school.name if school else 'the school'}'s parent portal.\n\n"
        f"Username: {username}\n"
        f"Temporary Password: {temp_password}\n\n"
        f"Please log in and change your password immediately. JazakAllahu Khairan."
    )
    whatsapp_number = "".join(ch for ch in (parent.phone or "") if ch.isdigit())
    if not whatsapp_number:
        flash("This parent has no phone number on file to message.", "warning")
        return redirect(url_for("parent_management.view_parent", parent_id=parent.id))

    return redirect(f"https://wa.me/{whatsapp_number}?text={quote(message)}")


# ---------------------------------------------------------------------------
# Parent detail + link/unlink students
# ---------------------------------------------------------------------------

@parent_management_bp.route("/<int:parent_id>")
@login_required
@roles_required(*ADMIN_ROLES)
def view_parent(parent_id):
    parent = Parent.query.get_or_404(parent_id)
    if not is_super_admin() and parent.school_id != current_school_id():
        abort(403)

    search = request.args.get("q", "").strip()
    search_results = []
    if search:
        like = f"%{search}%"
        query = Student.query.filter(Student.school_id == parent.school_id, Student.status == "active")
        query = query.filter(
            (Student.first_name.ilike(like)) | (Student.last_name.ilike(like)) | (Student.student_id.ilike(like))
        )
        already_linked_ids = {link.student_id for link in parent.parent_links}
        search_results = [s for s in query.limit(15).all() if s.id not in already_linked_ids]

    return render_template("parent_management/detail.html", parent=parent, search=search, search_results=search_results)


@parent_management_bp.route("/<int:parent_id>/link/<int:student_id>", methods=["POST"])
@login_required
@roles_required(*ADMIN_ROLES)
def link_student(parent_id, student_id):
    parent = Parent.query.get_or_404(parent_id)
    student = Student.query.get_or_404(student_id)
    if not is_super_admin() and parent.school_id != current_school_id():
        abort(403)
    if student.school_id != parent.school_id:
        flash("That student belongs to a different school and can't be linked.", "danger")
        return redirect(url_for("parent_management.view_parent", parent_id=parent.id))

    existing = ParentStudent.query.filter_by(parent_id=parent.id, student_id=student.id).first()
    if existing:
        flash(f"{student.full_name} is already linked to this parent.", "info")
        return redirect(url_for("parent_management.view_parent", parent_id=parent.id))

    has_existing_links = ParentStudent.query.filter_by(parent_id=parent.id).count() > 0
    relationship = request.form.get("relationship", parent.relationship or "Guardian")
    link = ParentStudent(
        parent_id=parent.id, student_id=student.id, relationship=relationship,
        is_primary=not has_existing_links,
    )
    db.session.add(link)
    if not parent.primary_student_id:
        parent.primary_student_id = student.id

    if safe_commit(log_context=f"link_student parent={parent_id} student={student_id}"):
        AuditLog.log(
            "settings_changed", description=f"{student.full_name} linked to parent {parent.full_name}",
            entity_type="parent", entity_id=parent.id, user=current_user,
        )
        flash(f"{student.full_name} linked to {parent.full_name}.", "success")
    return redirect(url_for("parent_management.view_parent", parent_id=parent.id))


@parent_management_bp.route("/<int:parent_id>/unlink/<int:student_id>", methods=["POST"])
@login_required
@roles_required(*ADMIN_ROLES)
def unlink_student(parent_id, student_id):
    parent = Parent.query.get_or_404(parent_id)
    if not is_super_admin() and parent.school_id != current_school_id():
        abort(403)

    link = ParentStudent.query.filter_by(parent_id=parent.id, student_id=student_id).first_or_404()
    student_name = link.student.full_name
    db.session.delete(link)  # only removes the relationship - never touches the Student record itself
    if parent.primary_student_id == student_id:
        parent.primary_student_id = None

    if safe_commit(log_context=f"unlink_student parent={parent_id} student={student_id}"):
        AuditLog.log(
            "settings_changed", description=f"{student_name} unlinked from parent {parent.full_name}",
            entity_type="parent", entity_id=parent.id, user=current_user,
        )
        flash(f"{student_name} unlinked from {parent.full_name}.", "info")
    return redirect(url_for("parent_management.view_parent", parent_id=parent.id))


@parent_management_bp.route("/<int:parent_id>/toggle-active", methods=["POST"])
@login_required
@roles_required(*ADMIN_ROLES)
def toggle_active(parent_id):
    parent = Parent.query.get_or_404(parent_id)
    if not is_super_admin() and parent.school_id != current_school_id():
        abort(403)
    parent.user.is_active_user = not parent.user.is_active_user
    if safe_commit(log_context=f"toggle_parent_active {parent_id}"):
        flash(f"Parent account {'activated' if parent.user.is_active_user else 'deactivated'}.", "info")
    return redirect(url_for("parent_management.view_parent", parent_id=parent.id))


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------

@parent_management_bp.route("/import", methods=["GET", "POST"])
@login_required
@roles_required(*ADMIN_ROLES)
def import_parents():
    school_id = current_school_id()
    if school_id is None:
        flash("Select a school first.", "warning")
        return redirect(url_for("parent_management.dashboard"))

    form = ParentImportForm()
    results = None

    if form.validate_on_submit():
        wb = load_workbook(form.file.data, data_only=True)
        ws = wb.active
        header = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        required_cols = {"parent_name", "phone", "student_id"}
        if not required_cols.issubset(set(header)):
            flash("Excel file must contain at least 'Parent Name', 'Phone', and 'Student ID' columns.", "danger")
            return redirect(url_for("parent_management.import_parents"))

        col = {name: idx for idx, name in enumerate(header)}
        created, failed, duplicates, missing_students = [], [], [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[col.get("parent_name", 0)]:
                continue
            try:
                with db.session.begin_nested():
                    parent_name = str(row[col["parent_name"]]).strip()
                    phone = str(row[col["phone"]]).strip() if row[col.get("phone", 0)] else ""
                    email = str(row[col["email"]]).strip().lower() if "email" in col and row[col["email"]] else None
                    relationship = str(row[col["relationship"]]).strip() if "relationship" in col and row[col["relationship"]] else "Guardian"
                    student_id_str = str(row[col["student_id"]]).strip() if row[col.get("student_id", 0)] else ""

                    if not phone or len(phone) < 7:
                        raise ValueError(f"Row {row_num}: invalid or missing phone number")
                    if email and "@" not in email:
                        raise ValueError(f"Row {row_num}: invalid email address")

                    student = Student.query.filter_by(school_id=school_id, student_id=student_id_str).first()
                    if not student:
                        missing_students.append(f"Row {row_num}: student ID '{student_id_str}' not found")
                        raise ValueError("missing student")

                    if email and User.query.filter_by(email=email).first():
                        duplicates.append(f"Row {row_num}: a user with email {email} already exists")
                        raise ValueError("duplicate email")

                    names = parent_name.split(" ", 1)
                    first_name, last_name = names[0], (names[1] if len(names) > 1 else names[0])
                    username = generate_unique_username("parent")
                    temp_password = generate_temp_password()

                    user = User(
                        username=username, email=email or f"{username}@placeholder.local",
                        first_name=first_name, last_name=last_name, phone=phone,
                        role=Role.PARENT, school_id=school_id, is_active_user=True, must_change_password=True,
                    )
                    user.set_password(temp_password)
                    db.session.add(user)
                    db.session.flush()

                    parent = Parent(
                        school_id=school_id, user_id=user.id, relationship=relationship,
                        primary_student_id=student.id,
                    )
                    db.session.add(parent)
                    db.session.flush()

                    db.session.add(ParentStudent(parent_id=parent.id, student_id=student.id, relationship=relationship, is_primary=True))
                    db.session.flush()

                created.append(f"Row {row_num}: {parent_name} linked to {student_id_str}")
            except ValueError:
                continue
            except Exception as exc:  # noqa: BLE001
                failed.append(f"Row {row_num}: {exc}")

        if safe_commit(log_context=f"import_parents school={school_id}"):
            AuditLog.log(
                "user_created", description=f"{len(created)} parents imported via Excel",
                user=current_user, school_id=school_id,
            )
            results = {
                "created": created, "failed": failed, "duplicates": duplicates, "missing_students": missing_students,
            }
            flash(f"{len(created)} parent accounts created.", "success")

    return render_template("parent_management/import.html", form=form, results=results)


@parent_management_bp.route("/import/template")
@login_required
@roles_required(*ADMIN_ROLES)
def download_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Parents"
    ws.append(["Parent Name", "Phone", "Email", "Relationship", "Student ID"])
    ws.append(["Abdul Rahman", "0240000000", "abdul@example.com", "Father", "MT001"])
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = 20

    import io
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem, as_attachment=True, download_name="parent_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
